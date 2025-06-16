'''
(C) John Bergqvist 2025

Script to extract the OD measurements from excel files

'''


# %%
import os
from openpyxl import load_workbook
import pandas as pd

#%%
def find_plate_origin(ws):
    """Finds the cell with '<>' in column A and returns start row and column"""
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value  # Column A = 1
        if isinstance(cell_value, str) and "<>" in cell_value:
            start_row = row + 1  # Data starts below '<>'
            start_col = 2        # Column B = 2 (to the right of '<>')
            return start_row, start_col
    raise ValueError("Couldn't find '<>' marker in column A.")

def process_excel(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    # Get plate name from G19 or fallback to G18
    plate_name = ws["G19"].value or ws["G18"].value
    if plate_name is None:
        raise ValueError(f"Plate name not found in G19 or G18 in file: {file_path}")

    # Locate OD600 data block
    start_row, start_col = find_plate_origin(ws)

    # Extract 8x12 well data
    data = []
    well_labels = []
    for i, row_letter in enumerate("ABCDEFGH"):
        row_data = []
        for col in range(12):  # 12 columns
            cell = ws.cell(row=start_row + i, column=start_col + col)
            row_data.append(cell.value)
            well_labels.append(f"{row_letter}{col + 1}")
        data.append(row_data)

    # Flatten the matrix and build DataFrame
    flat_data = [val for sublist in data for val in sublist]
    df = pd.DataFrame({
        "Well": well_labels,
        "OD600": flat_data,
        "plate_name": plate_name
    })

    return df

def process_run_folder(run_path):
    dfs = []
    for filename in os.listdir(run_path):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(run_path, filename)
            try:
                df = process_excel(file_path)
                dfs.append(df)
            except Exception as e:
                print(f"Error processing {file_path}: {e}")

    if dfs:
        combined_df = pd.concat(dfs, ignore_index=True)
        output_file = os.path.join(run_path, os.path.basename(run_path) + "_OD600_combined.csv")
        combined_df.to_csv(output_file, index=False)
        print(f"Saved combined CSV: {output_file}")
    else:
        print(f"No data found in {run_path}")

# Root directory where Run folders are located
root_dir = "/Volumes/behavgenom$/John/data_exp_info/BT/FoldSeek/commercial_proteins/experiments/MoilityAssay/AuxiliaryFiles/OD"

# Traverse Run folders
for folder in os.listdir(root_dir):
    if folder.startswith("Run") and os.path.isdir(os.path.join(root_dir, folder)):
        process_run_folder(os.path.join(root_dir, folder))



# %%
# Create a column called 'date_yyyymmdd' from the data in the column 'plate_name' (20250418_FS5_Cult_1_1_A_1)
def extract_date_from_plate_name(df):
    if 'plate_name' in df.columns:
        df['plate_name'] = df['plate_name'].astype(str)  # Ensure column values are strings
        df['date_yyyymmdd'] = df['plate_name'].str.extract(r'(\d{8})')[0]  # Extract date
        df['plate_name'] = df['plate_name'].str.replace(r'\d{8}', '', regex=True).str.strip()  # Remove date from plate_name
        df['plate_name'] = df['plate_name'].str.strip('_')  # Remove leading/trailing underscores
    else:
        print("Column 'plate_name' not found in the DataFrame.")
    return df

# process each csv file in the Run folders
def process_csv_files_in_run_folder(run_path):
    for filename in os.listdir(run_path):
        if filename.endswith("_OD600_combined.csv"):
            file_path = os.path.join(run_path, filename)
            df = pd.read_csv(file_path)
            df = extract_date_from_plate_name(df)
            df.to_csv(file_path, index=False)
            print(f"Updated date in {file_path}")

# Process each Run folder to update the date column
for folder in os.listdir(root_dir):
    if folder.startswith("Run") and os.path.isdir(os.path.join(root_dir, folder)):
        process_csv_files_in_run_folder(os.path.join(root_dir, folder))

# %%

# Combine all the 'Run' csv files into one csv file where the folder run name is added to a column called 'Run'
def combine_run_csv_files(root_dir):
    combined_data = []
    for folder in os.listdir(root_dir):
        if folder.startswith("Run") and os.path.isdir(os.path.join(root_dir, folder)):
            run_path = os.path.join(root_dir, folder)
            for filename in os.listdir(run_path):
                if filename.endswith("_OD600_combined.csv"):
                    file_path = os.path.join(run_path, filename)
                    df = pd.read_csv(file_path)
                    df['Run'] = folder  # Add Run column
                    combined_data.append(df)

    if combined_data:
        combined_df = pd.concat(combined_data, ignore_index=True)
        output_file = os.path.join(root_dir, "combined_OD600_data.csv")
        combined_df.to_csv(output_file, index=False)
        print(f"Saved combined CSV: {output_file}")
    else:
        print("No data found to combine.")
# Combine all Run csv files into one
combine_run_csv_files(root_dir)
# %%
